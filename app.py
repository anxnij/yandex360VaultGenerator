# Программа собирает свод.xlsx из самой свежей .xlsx-книги в публичной папке Яндекс.Диска 360,
# добавляя только новые строки и сохраняя форматирование итогового файла.
import os
import sys
import json
import re
import tempfile
import subprocess
import threading
import shutil
import hashlib
from pathlib import Path
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from time import perf_counter
from copy import copy as xlcopy

# Подключаем системное хранилище сертификатов Windows, чтобы HTTPS-запросы были стабильнее на пользовательских ПК.
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    import truststore
    truststore.inject_into_ssl()
except Exception:
    pass

# Сторонние библиотеки проекта: запросы к API, обработка Excel и графический интерфейс.
import requests
import pandas as pd
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
CONFIG_PATH = BASE_DIR / "config.json"

def load_app_config() -> dict:
    default_config = {
        "YANDEX_PUBLIC_LINK_DEFAULT": "",
        "YANDEX_API": "",
    }

    if CONFIG_PATH.exists():
        try:
            user_config = json.loads(CONFIG_PATH.read_text("utf-8"))
            default_config.update(user_config)
        except Exception:
            pass

    return default_config

APP_CONFIG = load_app_config()
YANDEX_PUBLIC_LINK_DEFAULT = APP_CONFIG["YANDEX_PUBLIC_LINK_DEFAULT"]
OUT_NAME_DEFAULT = "свод.xlsx"
TASK_NAME_DEFAULT = "Yandex360_Svod_Daily"
YANDEX_API = APP_CONFIG["YANDEX_API"]

# Рабочая папка приложения в AppData. Здесь лежат настройки, токен, лог последнего запуска и файлы ключей.
APP_DIR = Path(os.environ.get("APPDATA", str(Path.home()))) / "YandexSvodApp"
APP_DIR.mkdir(parents=True, exist_ok=True)

SETTINGS_PATH = APP_DIR / "settings.json"
TOKEN_PATH = APP_DIR / "token.dpapi"
LAST_RUN_LOG = APP_DIR / "last_run.log"

HTTP_TIMEOUT_LIST = 60
HTTP_TIMEOUT_DOWNLOAD = 300

# Названия колонок и значения фильтров, которые используются при очистке новых строк.
GRUZIM_COL_NAME = "Грузим?"
DIFF_COL_NAME = "Разница за штуку"
DATE_COL_NAME = "Requested deliv.date"

GRUZIM_SKIP_STATUSES = {
    "оставить блок на заказе будет доп инфа",
    "нет",
    "нет, информационный",
    "не заполнено",
}

DIFF_EPS = Decimal("0.2")

# Шифрует токен средствами Windows DPAPI, чтобы не хранить его в открытом виде на диске.
def dpapi_encrypt(plain: str) -> bytes:
    try:
        import win32crypt
    except ImportError:
        raise RuntimeError("Нужен пакет pywin32: pip install pywin32")
    return win32crypt.CryptProtectData(plain.encode("utf-8"), None, None, None, None, 0)


# Расшифровывает токен, ранее сохранённый через DPAPI.
def dpapi_decrypt(blob: bytes) -> str:
    try:
        import win32crypt
    except ImportError:
        raise RuntimeError("Нужен пакет pywin32: pip install pywin32")
    return win32crypt.CryptUnprotectData(blob, None, None, None, 0)[1].decode("utf-8")


# Сохраняет OAuth-токен в защищённый файл token.dpapi.
def save_token_secure(token: str) -> None:
    TOKEN_PATH.write_bytes(dpapi_encrypt(token.strip()))


# Загружает и расшифровывает токен из token.dpapi. Если что-то пошло не так — возвращает пустую строку.
def load_token_secure() -> str:
    if not TOKEN_PATH.exists():
        return ""
    try:
        return dpapi_decrypt(TOKEN_PATH.read_bytes()).strip()
    except Exception:
        return ""


# Загружает пользовательские настройки из settings.json. Если файла нет или он повреждён — отдаёт значения по умолчанию.
def load_settings() -> dict:
    if SETTINGS_PATH.exists():
        try:
            return json.loads(SETTINGS_PATH.read_text("utf-8"))
        except Exception:
            pass
    return {
        "public_link": YANDEX_PUBLIC_LINK_DEFAULT,
        "out_dir": str(Path.home()),
        "task_name": TASK_NAME_DEFAULT,
        "task_time": "09:00",
        "task_enabled": False,
    }


# Сохраняет текущие настройки интерфейса и автозапуска в settings.json.
def save_settings(s: dict) -> None:
    SETTINGS_PATH.write_text(json.dumps(s, ensure_ascii=False, indent=2), "utf-8")


# Возвращает текущий год по локальному времени.
def now_year() -> int:
    return datetime.now().year

# Нормализует имя колонки: убирает переносы строк, лишние пробелы и приводит к удобному виду для сравнения.
def norm_col(x) -> str:
    if x is None:
        return ""
    s = str(x).replace("\n", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


# Нормализует текст и делает casefold для надёжного сравнения строк без учёта регистра.
def norm_text_casefold(x) -> str:
    if x is None:
        return ""
    s = str(x).replace("\n", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip().casefold()


# Пытается аккуратно превратить денежное значение в Decimal: понимает "руб", пробелы, запятые и лишние символы.
def parse_rub_decimal(v) -> Decimal | None:
    if v is None:
        return None

    try:
        if pd.isna(v):
            return None
    except Exception:
        pass

    s = str(v).strip()
    if not s:
        return None

    s = s.lower().replace("руб.", "").replace("руб", "")
    s = s.replace("\xa0", " ").replace(" ", "")
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)

    if s in ("", "-", ".", "-."):
        return None

    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


# Проверяет, попадает ли значение "Разница за штуку" в диапазон [-0.2; 0.2].
def in_small_diff_range(v) -> bool:
    d = parse_rub_decimal(v)
    return d is not None and abs(d) <= DIFF_EPS


# Извлекает год из разных форматов даты/строки. Функция осталась как универсальный хелпер на случай возврата фильтра по году.
def parse_year_robust(v) -> int | None:
    if v is None:
        return None

    if isinstance(v, datetime):
        return v.year

    try:
        if pd.isna(v):
            return None
    except Exception:
        pass

    s = str(v).strip()
    if not s:
        return None

    m = re.search(r"(19|20)\d{2}", s)
    if m:
        try:
            return int(m.group(0))
        except Exception:
            pass

    dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
    if pd.isna(dt):
        return None
    return int(dt.year)


# Нормализует значение ячейки для построения стабильного ключа строки: даты, числа и текст приводятся к единому формату.
def norm_val_for_key(v) -> str:
    if v is None:
        return ""

    try:
        if pd.isna(v):
            return ""
    except Exception:
        pass

    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")

    s = str(v).strip()
    if s == "":
        return ""

    compact = s.replace("\xa0", " ").replace(" ", "").replace(",", ".")
    if re.fullmatch(r"-?\d+(\.\d+)?", compact):
        if re.fullmatch(r"0\d+", s):
            return s
        try:
            d = Decimal(compact)
            if d == d.to_integral_value():
                return str(int(d))
            return format(d.normalize(), "f")
        except Exception:
            return s

    return s


# Строит SHA1-ключ строки. Этот ключ нужен, чтобы быстро находить дубли между запусками.
def row_key_sha1(values: list[str]) -> str:
    return hashlib.sha1("␟".join(values).encode("utf-8")).hexdigest()


# Формирует HTTP-заголовки с OAuth-токеном для запросов в API Яндекс.Диска.
def make_headers(token: str) -> dict:
    return {"Authorization": f"OAuth {token}"}


# Безопасно парсит ISO-дату из метаданных API. Если дата битая — возвращает минимальную.
def _parse_iso_dt(s: str) -> datetime:
    if not s:
        return datetime.min.replace(tzinfo=timezone.utc)
    s = s.strip()
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except Exception:
        return datetime.min.replace(tzinfo=timezone.utc)

# Для каждого файла свода создаёт отдельный файл с ключами строк, чтобы не смешивать данные разных сводов.
def keys_file_for_outpath(out_path: Path) -> Path:
    h = hashlib.sha1(str(out_path).encode("utf-8")).hexdigest()[:16]
    return APP_DIR / f"keys_{h}.txt"

# Загружает ранее сохранённые ключи строк из файла keys_*.txt.
def load_keys(out_path: Path) -> set[str]:
    fp = keys_file_for_outpath(out_path)
    if not fp.exists():
        return set()
    try:
        return {x.strip() for x in fp.read_text("utf-8").splitlines() if x.strip()}
    except Exception:
        return set()


# Полностью перезаписывает файл ключей на диске.
def save_keys(out_path: Path, keys: set[str]) -> None:
    fp = keys_file_for_outpath(out_path)
    fp.write_text("\n".join(sorted(keys)) + ("\n" if keys else ""), "utf-8")


# Дописывает только новые ключи в конец файла ключей.
def append_keys(out_path: Path, new_keys: list[str]) -> None:
    if not new_keys:
        return
    fp = keys_file_for_outpath(out_path)
    with fp.open("a", encoding="utf-8") as f:
        for k in new_keys:
            f.write(k + "\n")


# Получает список всех объектов в публичной папке Яндекс.Диска через API с постраничной загрузкой.
def list_public_folder_items(token: str, public_key: str) -> list[dict]:
    headers = make_headers(token)
    items = []
    offset = 0
    limit = 200

    while True:
        r = requests.get(
            f"{YANDEX_API}/public/resources",
            headers=headers,
            params={"public_key": public_key, "limit": limit, "offset": offset},
            timeout=HTTP_TIMEOUT_LIST
        )
        if r.status_code in (401, 403):
            raise PermissionError(f"{r.status_code}: Нет доступа. Проверьте токен и права.")
        r.raise_for_status()

        data = r.json()
        page_items = data.get("_embedded", {}).get("items", [])
        if not page_items:
            break

        items.extend(page_items)

        if len(page_items) < limit:
            break
        offset += limit

    return items


# Из всех файлов папки выбирает самый свежий .xlsx по modified/created.
def pick_latest_xlsx_item(items: list[dict]) -> dict:
    xlsx_items = [
        it for it in items
        if it.get("type") == "file" and str(it.get("name", "")).lower().endswith(".xlsx")
    ]
    if not xlsx_items:
        raise FileNotFoundError("В папке не найдено ни одного .xlsx")

    def item_key(it: dict):
        return _parse_iso_dt(it.get("modified") or it.get("created") or "")

    return max(xlsx_items, key=item_key)


# Запрашивает у API временную ссылку на скачивание выбранного файла.
def get_public_download_href(token: str, public_key: str, path_in_public: str) -> str:
    r = requests.get(
        f"{YANDEX_API}/public/resources/download",
        headers=make_headers(token),
        params={"public_key": public_key, "path": path_in_public},
        timeout=HTTP_TIMEOUT_LIST
    )
    if r.status_code in (401, 403):
        raise PermissionError(f"{r.status_code}: Нет доступа к скачиванию файла.")
    r.raise_for_status()
    href = r.json().get("href")
    if not href:
        raise RuntimeError("Не удалось получить ссылку на скачивание.")
    return href


# Скачивает файл по прямой ссылке потоково, чтобы не держать всё содержимое в памяти.
def download_file(href: str, dest_path: Path) -> None:
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    with requests.get(href, stream=True, timeout=HTTP_TIMEOUT_DOWNLOAD) as r:
        r.raise_for_status()
        with open(dest_path, "wb") as f:
            for chunk in r.iter_content(chunk_size=1024 * 512):
                if chunk:
                    f.write(chunk)


# Оставляет только первый лист в книге Excel. Все остальные листы удаляются.
def drop_extra_sheets_keep_first(wb) -> None:
    while len(wb.worksheets) > 1:
        wb.remove(wb.worksheets[-1])


# Читает заголовки из первой строки листа и нормализует их.
def get_ws_headers(ws, header_row: int = 1) -> list[str]:
    headers = []
    for c in range(1, ws.max_column + 1):
        headers.append(norm_col(ws.cell(header_row, c).value))
    while headers and headers[-1] == "":
        headers.pop()
    return headers

# Ищет последнюю реально заполненную строку на листе.
def find_last_data_row(ws) -> int:
    for r in range(ws.max_row, 0, -1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v not in (None, ""):
                return r
    return 1


# Ищет первую заполненную строку с данными, чтобы использовать её как шаблон форматирования для новых строк.
def find_template_row(ws) -> int:
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if ws.cell(r, c).value not in (None, ""):
                return r
    return 2


# Если файл ключей потерян, пересобирает его по уже существующим строкам листа.
def rebuild_keys_from_sheet(ws, headers: list[str], out_path: Path) -> set[str]:
    keys = set()
    max_col = len(headers)

    for r in range(2, ws.max_row + 1):
        vals = [norm_val_for_key(ws.cell(r, c).value) for c in range(1, max_col + 1)]
        if all(v == "" for v in vals):
            continue
        keys.add(row_key_sha1(vals))

    save_keys(out_path, keys)
    return keys


# Добавляет новые строки в конец листа и копирует форматирование из шаблонной строки.
def append_rows_preserve_format(ws, headers: list[str], new_df: pd.DataFrame) -> int:
    if new_df.empty:
        return 0

    for h in headers:
        if h not in new_df.columns:
            new_df[h] = ""

    new_df = new_df.reindex(columns=headers).fillna("")

    last_row = find_last_data_row(ws)
    template_row = last_row if last_row >= 2 else find_template_row(ws)
    start_row = last_row + 1

    template_cells = [ws.cell(template_row, c) for c in range(1, len(headers) + 1)]
    template_height = ws.row_dimensions[template_row].height

    added = 0

    for i, row in enumerate(new_df.itertuples(index=False, name=None)):
        target_row = start_row + i

        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(target_row, col_idx)
            tcell = template_cells[col_idx - 1]

            cell.value = None if val is None or str(val).strip() == "" else val
            cell.font = xlcopy(tcell.font)
            cell.fill = xlcopy(tcell.fill)
            cell.border = xlcopy(tcell.border)
            cell.alignment = xlcopy(tcell.alignment)
            cell.number_format = tcell.number_format
            cell.protection = xlcopy(tcell.protection)

        if template_height is not None:
            ws.row_dimensions[target_row].height = template_height

        added += 1

    return added


# Читает первый лист свежего Excel-файла в DataFrame, нормализует заголовки и очищает пустые значения.
def read_latest_df(xlsx_path: Path) -> pd.DataFrame:
    df = pd.read_excel(xlsx_path, sheet_name=0, engine="openpyxl", dtype=str)
    df = df.rename(columns=lambda c: norm_col(c))
    df = df.dropna(axis=1, how="all")
    df = df.fillna("")
    return df


# Применяет фильтры только к добавляемым строкам: пустые, стоп-статусы в "Грузим?", диапазон по "Разница за штуку".
def filter_latest_df_only(df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    stats = {
        "skip_empty": 0,
        "skip_gruzim": 0,
        "skip_diff": 0,
        "skip_year": 0,
    }

    if df.empty:
        return df, stats

    empty_mask = df.astype(str).apply(lambda col: col.str.strip()).eq("").all(axis=1)
    stats["skip_empty"] = int(empty_mask.sum())
    df = df.loc[~empty_mask].copy()

    if df.empty:
        return df, stats

    gr_col = norm_col(GRUZIM_COL_NAME)
    if gr_col in df.columns:
        gr_mask = df[gr_col].map(norm_text_casefold).isin(GRUZIM_SKIP_STATUSES)
        stats["skip_gruzim"] = int(gr_mask.sum())
        df = df.loc[~gr_mask].copy()

    if df.empty:
        return df, stats

    diff_col = norm_col(DIFF_COL_NAME)
    if diff_col in df.columns:
        diff_mask = df[diff_col].map(in_small_diff_range)
        stats["skip_diff"] = int(diff_mask.sum())
        df = df.loc[~diff_mask].copy()

    if df.empty:
        return df, stats

    stats["skip_year"] = 0

    return df, stats

# Маленький помощник для пошагового логирования в UI и CLI.
class StepReporter:
    # Инициализация объекта.
    def __init__(self, callback=None):
        self.callback = callback

    # Отправляет произвольное сообщение в callback, если он передан.
    def log(self, message: str):
        if self.callback:
            self.callback(message)

    # Логирует завершение конкретного шага и время его выполнения.
    def step_done(self, title: str, seconds: float):
        self.log(f"{title} — готово за {seconds:.2f} сек.")


# Главный сценарий обновления: найти свежий файл, скачать, отфильтровать строки, убрать дубли, добавить новые строки в свод и сохранить результат.
def update_server_svod_from_latest(
    token: str,
    public_key: str,
    out_dir: Path,
    out_name: str,
    status_callback=None
) -> tuple[Path, int, str, dict]:
    rep = StepReporter(status_callback)
    total_started = perf_counter()

    stats = {
        "latest_rows_total": 0,
        "skip_empty": 0,
        "skip_gruzim": 0,
        "skip_diff": 0,
        "skip_year": 0,
        "latest_rows_after_filters": 0,
        "latest_rows_after_inner_dedup": 0,
        "new_rows_found": 0,
        "added_new_rows": 0,
    }

    t = perf_counter()
    items = list_public_folder_items(token, public_key)
    latest_item = pick_latest_xlsx_item(items)
    latest_name = latest_item.get("name") or "latest.xlsx"
    latest_path_in_public = latest_item.get("path")
    if not latest_path_in_public:
        raise RuntimeError("У свежего файла нет path в метаданных.")
    rep.log(f"1) Найден самый свежий файл: {latest_name}")
    rep.step_done("Поиск свежего файла", perf_counter() - t)

    t = perf_counter()
    tmp_dir = Path(tempfile.mkdtemp(prefix="yadisk_latest_"))
    local_latest = tmp_dir / latest_name
    href = get_public_download_href(token, public_key, latest_path_in_public)
    download_file(href, local_latest)
    rep.step_done("Скачивание свежего файла", perf_counter() - t)

    out_path = out_dir / out_name
    out_dir.mkdir(parents=True, exist_ok=True)

    if not out_path.exists():
        t = perf_counter()
        shutil.copyfile(local_latest, out_path)
        wb_new = load_workbook(out_path)
        drop_extra_sheets_keep_first(wb_new)
        wb_new.save(out_path)
        wb_new.close()
        rep.step_done("Создание нового свода из свежего файла", perf_counter() - t)

    t = perf_counter()
    wb = load_workbook(out_path)
    drop_extra_sheets_keep_first(wb)
    ws = wb.worksheets[0]

    headers = get_ws_headers(ws, 1)
    if not headers:
        wb.close()
        raise RuntimeError("В своде не найдены заголовки в первой строке.")

    existing_keys = load_keys(out_path)
    if not existing_keys:
        existing_keys = rebuild_keys_from_sheet(ws, headers, out_path)

    rep.step_done("Подготовка свода и ключей", perf_counter() - t)

    t = perf_counter()
    latest_df = read_latest_df(local_latest)
    stats["latest_rows_total"] = int(len(latest_df))
    rep.step_done("Чтение свежего файла", perf_counter() - t)

    t = perf_counter()
    for h in headers:
        if h not in latest_df.columns:
            latest_df[h] = ""

    latest_df = latest_df.reindex(columns=headers).fillna("")
    latest_df, fstats = filter_latest_df_only(latest_df)
    stats.update(fstats)
    stats["latest_rows_after_filters"] = int(len(latest_df))
    rep.step_done("Фильтрация добавляемых строк", perf_counter() - t)

    t = perf_counter()
    if not latest_df.empty:
        latest_df = latest_df.drop_duplicates(subset=headers, keep="first")
    stats["latest_rows_after_inner_dedup"] = int(len(latest_df))
    rep.step_done("Удаление дублей внутри свежего файла", perf_counter() - t)

    t = perf_counter()
    new_rows = []
    new_keys = []

    if not latest_df.empty:
        for row in latest_df.itertuples(index=False, name=None):
            vals = [norm_val_for_key(v) for v in row]
            k = row_key_sha1(vals)
            if k not in existing_keys:
                new_rows.append(row)
                new_keys.append(k)
                existing_keys.add(k)

    stats["new_rows_found"] = len(new_rows)
    rep.step_done("Поиск новых строк", perf_counter() - t)

    t = perf_counter()
    added = 0
    if new_rows:
        new_df = pd.DataFrame(new_rows, columns=headers)
        added = append_rows_preserve_format(ws, headers, new_df)
        append_keys(out_path, new_keys)

    stats["added_new_rows"] = added
    rep.step_done("Добавление новых строк в свод", perf_counter() - t)

    t = perf_counter()
    drop_extra_sheets_keep_first(wb)
    wb.save(out_path)
    wb.close()
    rep.step_done("Сохранение свода", perf_counter() - t)

    total_seconds = perf_counter() - total_started
    rep.log(f"Итог: обновление завершено за {total_seconds:.2f} сек.")
    rep.log(f"Добавлено новых строк: {added}")

    try:
        shutil.rmtree(tmp_dir, ignore_errors=True)
    except Exception:
        pass

    return out_path, added, latest_name, stats


# Проверяет, что время введено в формате HH:MM.
def validate_hhmm(s: str) -> str:
    try:
        dt = datetime.strptime(s.strip(), "%H:%M")
        return dt.strftime("%H:%M")
    except Exception:
        raise ValueError("Время должно быть в формате HH:MM, например 09:00")


# Возвращает путь к текущему исполняемому файлу/интерпретатору. Используется для автозапуска через Планировщик задач Windows.
def exe_path() -> str:
    return sys.executable

# Собирает аргументы командной строки для фонового запуска обновления.
def script_args_for_task(public_link: str, out_dir: str) -> str:
    return f'--run-once --link="{public_link}" --outdir="{out_dir}"'


# Создаёт или обновляет ежедневную задачу Windows Task Scheduler.
def create_or_update_task(task_name: str, time_hhmm: str, public_link: str, out_dir: str) -> None:
    time_hhmm = validate_hhmm(time_hhmm)
    tr = f'"{exe_path()}" {script_args_for_task(public_link, out_dir)}'
    cmd = [
        "schtasks",
        "/Create",
        "/F",
        "/SC", "DAILY",
        "/TN", task_name,
        "/TR", tr,
        "/ST", time_hhmm
    ]
    p = subprocess.run(cmd, capture_output=True, text=True, shell=False)
    if p.returncode != 0:
        raise RuntimeError(f"Не удалось создать или обновить задачу.\n\n{p.stdout}\n{p.stderr}")


# Удаляет задачу автозапуска из Планировщика задач Windows.
def delete_task(task_name: str) -> None:
    subprocess.run(
        ["schtasks", "/Delete", "/F", "/TN", task_name],
        capture_output=True,
        text=True,
        shell=False
    )


# Главное окно приложения на Tkinter: хранит состояние интерфейса и связывает UI с бизнес-логикой.
class App(tk.Tk):
    # Инициализация объекта.
    def __init__(self):
        super().__init__()
        self.title("Свод из Яндекс.Диска 360")
        self.geometry("900x560")

        self.settings = load_settings()

        self.var_link = tk.StringVar(value=self.settings.get("public_link", YANDEX_PUBLIC_LINK_DEFAULT))
        self.var_out_dir = tk.StringVar(value=self.settings.get("out_dir", str(Path.home())))
        self.var_time = tk.StringVar(value=self.settings.get("task_time", "09:00"))
        self.var_task_name = tk.StringVar(value=self.settings.get("task_name", TASK_NAME_DEFAULT))
        self.var_status = tk.StringVar(value="")

        token_exists = bool(load_token_secure())
        self.var_token_state = tk.StringVar(value=("✅ токен сохранён" if token_exists else "⚠️ токен не задан"))

        self._build()

    # Собирает все элементы интерфейса: поля, кнопки, блок статуса.
    def _build(self):
        pad = {"padx": 10, "pady": 6}

        frm = ttk.Frame(self)
        frm.pack(fill="both", expand=True, padx=12, pady=12)

        ttk.Label(frm, text="Ссылка на папку Яндекс.Диска:").grid(row=0, column=0, sticky="w", **pad)
        ttk.Entry(frm, textvariable=self.var_link, width=95).grid(row=1, column=0, columnspan=3, sticky="we", **pad)

        ttk.Label(frm, text="Папка, где лежит свод.xlsx:").grid(row=2, column=0, sticky="w", **pad)
        ttk.Entry(frm, textvariable=self.var_out_dir, width=75).grid(row=3, column=0, columnspan=2, sticky="we", **pad)
        ttk.Button(frm, text="Выберите папку", command=self.pick_out_dir).grid(row=3, column=2, sticky="e", **pad)

        ttk.Label(frm, text="Токен доступа (OAuth y0__...):").grid(row=4, column=0, sticky="w", **pad)
        ttk.Label(frm, textvariable=self.var_token_state).grid(row=4, column=1, sticky="w", **pad)
        ttk.Button(frm, text="Введите или обновите токен", command=self.set_token_dialog).grid(row=4, column=2, sticky="e", **pad)

        ttk.Separator(frm).grid(row=5, column=0, columnspan=3, sticky="we", pady=10)

        ttk.Label(frm, text="Автозапуск ежедневно:").grid(row=6, column=0, sticky="w", **pad)
        ttk.Label(frm, text="Имя задачи:").grid(row=7, column=0, sticky="w", **pad)
        ttk.Entry(frm, textvariable=self.var_task_name, width=35).grid(row=7, column=1, sticky="w", **pad)

        ttk.Label(frm, text="Время (HH:MM):").grid(row=7, column=2, sticky="w", **pad)
        ttk.Entry(frm, textvariable=self.var_time, width=10).grid(row=8, column=2, sticky="w", **pad)

        btns = ttk.Frame(frm)
        btns.grid(row=8, column=0, columnspan=2, sticky="w", padx=10, pady=6)
        ttk.Button(btns, text="Обновить свод", command=self.run_now).pack(side="left", padx=6)
        ttk.Button(btns, text="Включить автозапуск", command=self.enable_task).pack(side="left", padx=6)
        ttk.Button(btns, text="Выключить автозапуск", command=self.disable_task).pack(side="left", padx=6)

        ttk.Separator(frm).grid(row=9, column=0, columnspan=3, sticky="we", pady=10)

        ttk.Label(frm, text="Статус:").grid(row=10, column=0, sticky="w", **pad)

        self.status_box = tk.Text(frm, height=16, wrap="word")
        self.status_box.grid(row=11, column=0, columnspan=3, sticky="nsew", padx=10, pady=6)
        self.status_box.configure(state="disabled")
        frm.columnconfigure(0, weight=1)
        frm.columnconfigure(1, weight=1)
        frm.rowconfigure(11, weight=1)

    # Полностью заменяет содержимое поля статуса.
    def set_status(self, text: str):
        self.status_box.configure(state="normal")
        self.status_box.delete("1.0", "end")
        self.status_box.insert("end", text)
        self.status_box.configure(state="disabled")
        self.update_idletasks()

    # Добавляет новую строку в поле статуса без очистки предыдущих сообщений.
    def add_status(self, text: str):
        self.status_box.configure(state="normal")
        if self.status_box.index("end-1c") != "1.0":
            self.status_box.insert("end", "\n")
        self.status_box.insert("end", text)
        self.status_box.see("end")
        self.status_box.configure(state="disabled")
        self.update_idletasks()

    # Открывает диалог выбора папки для итогового свода.
    def pick_out_dir(self):
        d = filedialog.askdirectory(initialdir=self.var_out_dir.get() or str(Path.home()))
        if d:
            self.var_out_dir.set(d)

    # Открывает отдельное окно, в котором пользователь вводит и сохраняет OAuth-токен.
    def set_token_dialog(self):
        win = tk.Toplevel(self)
        win.title("Введите OAuth токен")
        win.geometry("680x220")

        ttk.Label(
            win,
            text="Вставьте OAuth токен (y0__...) — он сохранится безопасно на этом ПК:"
        ).pack(padx=10, pady=10, anchor="w")

        txt = tk.Text(win, height=3, width=82)
        txt.pack(padx=10, pady=6)

        # Вложенный обработчик кнопки "Сохранить" внутри окна ввода токена.
        def save():
            token = txt.get("1.0", "end").strip()
            if not token.startswith("y0__"):
                messagebox.showerror("Ошибка", "Похоже, токен неверный. Обычно он начинается с y0__.")
                return
            try:
                save_token_secure(token)
                self.var_token_state.set("✅ токен сохранён")
                win.destroy()
            except Exception as e:
                messagebox.showerror("Ошибка", str(e))

        ttk.Button(win, text="Сохранить", command=save).pack(padx=10, pady=10, anchor="e")

    # Сохраняет актуальные значения из интерфейса в settings.json.
    def persist_settings(self, task_enabled: bool | None = None):
        s = {
            "public_link": self.var_link.get().strip(),
            "out_dir": self.var_out_dir.get().strip(),
            "task_name": self.var_task_name.get().strip() or TASK_NAME_DEFAULT,
            "task_time": self.var_time.get().strip() or "09:00",
            "task_enabled": self.settings.get("task_enabled", False) if task_enabled is None else task_enabled,
        }
        self.settings = s
        save_settings(s)

    # Запускает ручное обновление в отдельном потоке, чтобы интерфейс не зависал.
    def run_now(self):
        threading.Thread(target=self._run_now_worker, daemon=True).start()

    # Рабочий поток ручного обновления: валидирует данные, запускает основной сценарий и обновляет UI.
    def _run_now_worker(self):
        try:
            self.persist_settings()

            token = load_token_secure()
            if not token:
                self.after(0, lambda: messagebox.showwarning(
                    "Нужен токен",
                    "Сначала нажмите «Введите или обновите токен» и вставьте токен."
                ))
                return

            out_dir = Path(self.var_out_dir.get()).expanduser().resolve()
            out_dir.mkdir(parents=True, exist_ok=True)

            self.after(0, lambda: self.set_status("Запуск обновления..."))

            # Вложенный callback для безопаского добавления статусов из фонового потока в UI.
            def ui_status(msg: str):
                self.after(0, lambda m=msg: self.add_status(m))

            out_path, added, latest_name, stats = update_server_svod_from_latest(
                token=token,
                public_key=self.var_link.get().strip(),
                out_dir=out_dir,
                out_name=OUT_NAME_DEFAULT,
                status_callback=ui_status
            )

            self.after(0, lambda: self.add_status(f"Всего строк в свежем файле: {stats['latest_rows_total']}"))
            self.after(0, lambda: self.add_status(f"Исключено пустых строк: {stats['skip_empty']}"))
            self.after(0, lambda: self.add_status(f"Исключено по '{GRUZIM_COL_NAME}': {stats['skip_gruzim']}"))
            self.after(0, lambda: self.add_status(f"Исключено по '{DIFF_COL_NAME}': {stats['skip_diff']}"))
            self.after(0, lambda: self.add_status(f"Осталось после фильтров: {stats['latest_rows_after_filters']}"))
            self.after(0, lambda: self.add_status(f"После удаления дублей внутри свежего файла: {stats['latest_rows_after_inner_dedup']}"))
            self.after(0, lambda: self.add_status(f"Найдено новых строк: {stats['new_rows_found']}"))
            self.after(0, lambda: self.add_status(f"Добавлено новых строк: {stats['added_new_rows']}"))
            self.after(0, lambda: self.add_status(f"Свод сохранён: {out_path}"))

            msg = (
                f"Готово ✅\n"
                f"Найден свежий файл: {latest_name}\n"
                f"Добавлено новых строк: {added}\n"
                f"Свод: {out_path}"
            )
            self.after(0, lambda: messagebox.showinfo("Готово", msg))

        except Exception as e:
            self.after(0, lambda: self.add_status(f"Ошибка: {e}"))
            self.after(0, lambda: messagebox.showerror("Ошибка", str(e)))

    # Включает ежедневный автозапуск через Планировщик задач Windows.
    def enable_task(self):
        try:
            self.persist_settings(task_enabled=True)

            token = load_token_secure()
            if not token:
                messagebox.showwarning("Нужен токен", "Сначала сохраните токен.")
                return

            task_name = self.var_task_name.get().strip() or TASK_NAME_DEFAULT
            time_hhmm = validate_hhmm(self.var_time.get().strip())

            create_or_update_task(
                task_name=task_name,
                time_hhmm=time_hhmm,
                public_link=self.var_link.get().strip(),
                out_dir=self.var_out_dir.get().strip()
            )

            msg = f"Автозапуск включён ✅ Каждый день в {time_hhmm} (задача: {task_name})"
            self.set_status(msg)
            messagebox.showinfo("Ок", msg)

        except Exception as e:
            self.set_status(f"Ошибка автозапуска: {e}")
            messagebox.showerror("Ошибка", str(e))

    # Выключает автозапуск, удаляя задачу из Планировщика задач.
    def disable_task(self):
        try:
            task_name = self.var_task_name.get().strip() or TASK_NAME_DEFAULT
            delete_task(task_name)
            self.persist_settings(task_enabled=False)

            msg = f"Автозапуск выключен ✅ (задача {task_name} удалена)"
            self.set_status(msg)
            messagebox.showinfo("Ок", msg)

        except Exception as e:
            self.set_status(f"Ошибка: {e}")
            messagebox.showerror("Ошибка", str(e))


# Режим запуска из командной строки. Используется автозапуском и может применяться без GUI.
def run_once_cli():
    import argparse

    ap = argparse.ArgumentParser()
    ap.add_argument("--run-once", action="store_true")
    ap.add_argument("--link", required=True)
    ap.add_argument("--outdir", required=True)
    args = ap.parse_args()

    token = load_token_secure()
    if not token:
        raise RuntimeError("Токен не сохранён. Откройте программу и вставьте токен один раз.")

    out_dir = Path(args.outdir).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    logs = []

    # Вложенный логгер для консольного режима.
    def cli_status(msg: str):
        logs.append(msg)
        print(msg)

    out_path, added, latest_name, stats = update_server_svod_from_latest(
        token=token,
        public_key=args.link.strip(),
        out_dir=out_dir,
        out_name=OUT_NAME_DEFAULT,
        status_callback=cli_status
    )

    logs.extend([
        f"Всего строк в свежем файле: {stats['latest_rows_total']}",
        f"Исключено пустых строк: {stats['skip_empty']}",
        f"Исключено по '{GRUZIM_COL_NAME}': {stats['skip_gruzim']}",
        f"Исключено по '{DIFF_COL_NAME}': {stats['skip_diff']}",
        f"Осталось после фильтров: {stats['latest_rows_after_filters']}",
        f"После удаления дублей внутри свежего файла: {stats['latest_rows_after_inner_dedup']}",
        f"Найдено новых строк: {stats['new_rows_found']}",
        f"Добавлено новых строк: {stats['added_new_rows']}",
        f"Свод: {out_path}",
        f"Свежий файл: {latest_name}",
    ])

    LAST_RUN_LOG.write_text(
        f"{datetime.now().isoformat()}\n" + "\n".join(logs) + "\n",
        "utf-8"
    )


if __name__ == "__main__":
    if "--run-once" in sys.argv:
        run_once_cli()
    else:
        app = App()
        app.mainloop()
